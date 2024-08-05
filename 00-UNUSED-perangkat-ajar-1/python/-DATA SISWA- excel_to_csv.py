import openpyxl
import os

listStr=""
directory="E:\\binaracad\\fsw-gc-adha\\perangkat_ajar\\data\\"
for file in os.listdir(directory):
    if ".xlsx" in file:
        WB = openpyxl.load_workbook(directory+file)
        WS = WB[WB.sheetnames[0]]

        maxRow = WS.max_row+1
        maxCol = WS.max_column+1

        WB.save(directory+file)
        WB.close()



for file in os.listdir(directory):
    if ".xlsx" in file:
        WB = openpyxl.load_workbook(directory+file)
        WS = WB[WB.sheetnames[0]]
        maxRow = WS.max_row+1
        maxCol = WS.max_column+1
        filename=file
        filename=filename.replace(".xlsx", "")
        csvFilename="DATA_SISWA.csv"
        with open(csvFilename, "w") as f:
            f.write("")
        for row in range(1, maxRow):
            for column in range(1, maxCol):
                valRow_1=WS.cell(row=row, column=column).value
                if valRow_1==None:
                    valRow_1=""
                valStr = str(valRow_1)
                valStr+=";"
                listStr+=valStr

            listStr=listStr[0:len(listStr)-1]
            with open(csvFilename, "a") as f:
                f.write(listStr)
                f.write("\n")
            listStr=""